import csv
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime


def _build_rows(users):
    headers = ["ID", "Username", "First Name", "Last Name", "Has Avatar"]
    rows = []
    for u in users:
        rows.append([
            u.get("id", ""),
            u.get("username", ""),
            u.get("first_name", ""),
            u.get("last_name", ""),
            "Yes" if u.get("has_avatar") else "No"
        ])
    return headers, rows


def _write_xlsx(users, filepath):
    wb = Workbook()
    ws = wb.active
    ws.title = "Users"
    headers, rows = _build_rows(users)
    hf = Font(bold=True, color="FFFFFF", name="Arial")
    hfill = PatternFill("solid", start_color="2F75B6")
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = hf
        c.fill = hfill
        c.alignment = Alignment(horizontal="center")
    for row_data in rows:
        ws.append(row_data)
    for col, w in zip(["A","B","C","D","E"], [15, 22, 20, 20, 12]):
        ws.column_dimensions[col].width = w
    wb.save(filepath)


def _write_csv(users, filepath):
    headers, rows = _build_rows(users)
    with open(filepath, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(headers)
        w.writerows(rows)


def create_export_files(users, prefix="users"):
    Path("exports").mkdir(exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    xlsx_path = f"exports/{prefix}_{ts}.xlsx"
    csv_path = f"exports/{prefix}_{ts}.csv"
    _write_xlsx(users, xlsx_path)
    _write_csv(users, csv_path)
    return xlsx_path, csv_path